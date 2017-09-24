from openpyxl import load_workbook  # Openpyxl is a Python library for reading and writing Excel 2010 
import datetime

def repair_time(my_time):  # функция для исправления типа даты и времени
    if isinstance(my_time=my_time, datetime.datetime):  #isinstance() по сравнению с type() позволяет проверить данное на принадлежность хотя бы одному типу из кортежа, переданного в качестве второго аргумента
        my_time = my_time.time()

    if isinstance(my_time=my_time, datetime.time):
        return my_time
    else:
        return ""

def read_excel(filename):
    work_book = load_workbook(filename)  # передаем имя файла, которе нужно прочитать. с файлами работаем постранично
    work_sheet = work_book("название листа")  # обращаемся к вкладке по ее названию
    
    excel_data = list()  # явное создание list() считается более правильным хз почему, вместо []

    for row in range(6,work_sheet.max_row):  # проходимся с колонки 6 по 41
        excel_row = dict()
        if work_sheet.cell(row=row, column=1).value is None:  # чтобы не читать пустые строки. надо именно is None писать, а не == None
            break
        excel_row["time_start"] = work_sheet.cell(row=row, column=1).value
        excel_row["time_end"] = work_sheet.cell(row=row, column=2).value
        excel_row["title"] = work_sheet.cell(row=row, column=3).value  # и т.д.

        excel_data.append(excel_row)

    return excel_data

def show_excel_data(data): 
    print(data)

    pass

if __name__ == "__main__":
    excel_data = read_excel("имя файла")
    show_excel_data(excel_data)
