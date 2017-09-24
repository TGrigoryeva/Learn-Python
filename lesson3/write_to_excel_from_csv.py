import csv
from openpyxl import Workbook #Openpyxl is a Python library for reading and writing Excel 2010 

def read_csv(filename):
    data = []

    with open(filename, "r", encoding = "utf-8") as f: #добавить имя файла, еще м.б. при лишних спец символах - "utf-8-sig"
        fields = [] # добавить заголовки
        reader = csv.DictReader(f, fields, delimiter =";")
        for row in reader:
            data.append(row)

    return data

def excel_write(data):
    workbook = Workbook() #создаем новый эксель workbook и помещаем в переменную
    worksheet = workbook.active  #создаем вкладку, с которой будем работать и выбираем первую активную через active
    worksheet.title = "Посещаемость" # создали заголовок

    worksheet.cell(row=1,column=1).value = "Имя" # создали заголовки в ячейках вручную
    worksheet.cell(row=1,column=2).value = "Пришел?"
    worksheet.cell(row=1,column=3).value = "Род деятельности"
    worksheet.cell(row=1,column=4).value = "Компания"

    row = 2
    for item in data:
        worksheet.cell(row=row,column=1).value = item["name"] # по ключам заполняем каждую строку
        worksheet.cell(row=row,column=2).value = item["visited"]
        worksheet.cell(row=row,column=3).value = item["is working"]
        worksheet.cell(row=row,column=4).value = item["company"]
        row += 1


    workbook.save("participants".xlsx)

csv_data = read_csv("participants.csv") #Прочитали файл и положили в переменную
excel_write(csv_data)

